import openpyxl
from openpyxl.styles import (PatternFill, Border, Side, Alignment,
                             Font)


def get_sheet_data_and_headers(filename, sheet_index=0):
    # 默认取第一个sheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[sheet_index]
    headers = []
    data = []
    for row in ws.iter_rows(max_row=1, min_col=2):
        for cell in row:
            cell_value = cell.value
            if not isinstance(cell_value, str):
                return [], []
            headers.append(cell_value.strip())
    for row in ws.iter_rows(min_row=2, min_col=2):
        row_data = []
        for cell in row:
            cell_value = cell.value
            if cell_value is None:
                cell_value = ""
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
            row_data.append(cell_value)
        data.append(row_data)
    return headers, data


def get_sheet_names(filename):
    return openpyxl.load_workbook(filename).sheetnames


def set_ws_border(ws, start_row, start_col, end_row, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = Border(left=Side(border_style="thin", color="000000"),
                                                         right=Side(border_style="thin", color="000000"),
                                                         top=Side(border_style="thin", color="000000"),
                                                         bottom=Side(border_style="thin", color="000000"))


def export_u_s_table_widget(u_s_table_widget, u_s_table_data, filename):
    data = {}  # key是用车单位 value是[] 分别为年度总收入，年度总支出，年度结余
    for yongchedanwei, item in u_s_table_data.items():
        if yongchedanwei not in data:
            data[yongchedanwei] = [0, 0, 0]
        for month, income_expense_balance in item.items():
            data[yongchedanwei][0] += income_expense_balance[0]
            data[yongchedanwei][1] += income_expense_balance[1]
            data[yongchedanwei][2] += income_expense_balance[2]
    try:
        if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
            filename += ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # 先写入统计数据
        ws.cell(row=1, column=1).value = "用车单位"
        ws.cell(row=1, column=2).value = "年度总收入"
        ws.cell(row=1, column=3).value = "年度总支出"
        ws.cell(row=1, column=4).value = "年度结余"
        row = 2
        for yongchedanwei, item in data.items():
            ws.cell(row=row, column=1).value = yongchedanwei
            ws.cell(row=row, column=2).value = item[0]
            ws.cell(row=row, column=3).value = item[1]
            ws.cell(row=row, column=4).value = item[2]
            row += 1
        # 再写入表格数据
        row += 1
        ws.cell(row=row, column=1).value = "详细数据如下："
        row += 1
        ws.cell(row=row, column=1).value = "用车单位"
        ws.cell(row=row, column=2).value = "月份"
        ws.cell(row=row, column=3).value = "收入"
        ws.cell(row=row, column=4).value = "支出"
        ws.cell(row=row, column=5).value = "结余"
        row += 1
        for i in range(u_s_table_widget.rowCount()):
            for j in range(u_s_table_widget.columnCount()):
                item = u_s_table_widget.item(i, j)
                item_text = item.text() if item else ""
                background_color = item.background().color().name().strip("#").upper()
                foreground_color = item.foreground().color().name().strip("#").upper()
                if background_color == "000000":
                    background_color = "FFFFFF"
                ws.cell(row=row, column=j + 1).value = item_text
                ws.cell(row=row, column=j + 1).fill = PatternFill("solid", fgColor=background_color)
                ws.cell(row=row, column=j + 1).font = Font(color=foreground_color)
            row += 1
        set_ws_border(ws, 1, 1, ws.max_row, ws.max_column)
        wb.save(filename)
        return True, "1"
    except Exception as e:
        return False, str(e)


def export_table_widget(table_widget, filename, number_cols=None):
    try:
        if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
            filename += ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # 合并第一行的单元格
        # 在第一行居中写入标题"常熟市顺发汽车租赁有限公司"并加粗加黑
        max_col = table_widget.columnCount()
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1).value = "常熟市顺发汽车租赁有限公司"
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=1).font = Font(bold=True, color="000000")

        # 在第二行设置表头
        for col in range(table_widget.columnCount()):
            ws.cell(row=2, column=col + 1).value = table_widget.horizontalHeaderItem(col).text()

        number_cols = set(number_cols) if number_cols else set()
        row_excel_to_write = 0
        # 设置表格内容
        for row in range(table_widget.rowCount()):
            # 查看是否可视
            if table_widget.isRowHidden(row):
                continue
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                item_content = item.text() if item else ""
                if col in number_cols and item_content:
                    try:
                        item_content = float(item_content)
                        ws.cell(row=row_excel_to_write + 3, column=col + 1).data_type = "int"
                        ws.cell(row=row_excel_to_write + 3, column=col + 1).number_format = "0.00"
                    except Exception as e:
                        print(e)
                ws.cell(row=row_excel_to_write + 3, column=col + 1).value = item_content
                # color must be aRBG
                background_color = item.background().color().name().strip("#").upper()
                foreground_color = item.foreground().color().name().strip("#").upper()
                if background_color == "000000":
                    background_color = "FFFFFF"
                ws.cell(row=row_excel_to_write + 3, column=col + 1).fill = PatternFill("solid",
                                                                                       fgColor=background_color)
                ws.cell(row=row_excel_to_write + 3, column=col + 1).font = Font(color=foreground_color)
            row_excel_to_write += 1
        # 计算为数字列的总和
        for col in number_cols:
            header_desc = "总" + table_widget.horizontalHeaderItem(col).text() + "："
            total = 0
            for row in range(table_widget.rowCount()):
                if table_widget.isRowHidden(row):
                    continue
                item = table_widget.item(row, col)
                item_content = item.text() if item else ""
                try:
                    item_content = float(item_content)
                    total += item_content
                except Exception as e:
                    print(e)
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column - 1).value = header_desc
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column - 1).font = Font(bold=True)
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column).value = total
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column).font = Font(bold=True)
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column).number_format = "0.00"
            ws.cell(row=row_excel_to_write + 3, column=ws.max_column).data_type = "int"
            row_excel_to_write += 1
        set_ws_border(ws, 1, 1, ws.max_row, ws.max_column)
        wb.save(filename)
        return True, "1"
    except Exception as e:
        return False, str(e)


def export_table_widget_with_money(user_table_widget, filename, pre_desc, total_money):
    try:
        if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
            filename += ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # 第一行写入总金额
        ws.cell(row=1, column=1).value = pre_desc
        ws.cell(row=1, column=2).value = total_money
        # 设置表头
        for col in range(user_table_widget.columnCount()):
            ws.cell(row=2, column=col + 1).value = user_table_widget.horizontalHeaderItem(col).text()

        # 设置表格内容
        for row in range(user_table_widget.rowCount()):
            for col in range(user_table_widget.columnCount()):
                item = user_table_widget.item(row, col)
                item_text = item.text() if item else ""
                background_color = item.background().color().name().strip("#").upper()
                foreground_color = item.foreground().color().name().strip("#").upper()
                if background_color == "000000":
                    background_color = "FFFFFF"
                ws.cell(row=row + 3, column=col + 1).fill = PatternFill("solid", fgColor=background_color)
                ws.cell(row=row + 3, column=col + 1).value = item_text
                ws.cell(row=row + 3, column=col + 1).font = Font(color=foreground_color)
        set_ws_border(ws, 1, 1, ws.max_row, ws.max_column)
        wb.save(filename)
        return True, "1"
    except Exception as e:
        return False, str(e)
