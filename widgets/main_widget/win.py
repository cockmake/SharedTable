import datetime
import threading

import openpyxl
from PyQt5 import QtGui
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QDesktopWidget, QFileDialog, QApplication, QTableWidgetItem

from utils.excel import get_sheet_names, export_table_widget, export_table_widget_with_money, export_u_s_table_widget
from utils.qhttp import Http
from widgets.create_user_table_dialog.win import UserTableCreateDialog
from widgets.create_wage_table_dialog.win import WageTableCreateDialog
from widgets.custom_tool_box.win import FunctionToolBox
from widgets.f_s_dialog.win import FSDialog
from widgets.main_table_add_dialog.win import MainTableAddDialog
from widgets.main_table_search_widget.win import MainTableSearchWidget
from widgets.main_widget.main_widget import Ui_MainWindow
from widgets.u_s_dialog.win import USDialog


class MainWidget(QtWidgets.QMainWindow, Ui_MainWindow):
    add_completed = pyqtSignal()

    def __init__(self, parent=None):
        super(MainWidget, self).__init__(parent=parent)
        self.setupUi(self)
        self.init_font()
        self.open_table_main.triggered.connect(self.open_table_main_slot)
        self.main_table_add.clicked.connect(self.main_table_add_slot)
        self.main_table_search.clicked.connect(self.main_table_search_slot)
        self.main_table_del_btn.clicked.connect(self.main_table_del_btn_slot)
        self.create_user_table_btn.clicked.connect(self.create_user_table_btn_slot)
        self.create_wage_table_btn.clicked.connect(self.create_wage_table_btn_slot)
        self.auto_save_checkbox = QtWidgets.QCheckBox("明细表自动保存", parent=self.main_table)
        self.auto_save_checkbox.setChecked(True)
        self.menubar.setCornerWidget(self.auto_save_checkbox)
        self.auto_save_checkbox.stateChanged.connect(self.auto_save_checkbox_state_changed_slot)
        self.col_index = {}  # 默认为main_table的
        self.main_table_result.setColumnCount(self.main_table.columnCount())
        for i in range(self.main_table.columnCount()):
            header_item = self.main_table.horizontalHeaderItem(i)
            self.col_index[header_item.text()] = i
            self.main_table_result.setHorizontalHeaderItem(i, header_item.clone())
        self.user_table_col_index = {}
        for i in range(self.user_table.columnCount()):
            self.user_table_col_index[self.user_table.horizontalHeaderItem(i).text()] = i
        self.wage_table_col_index = {}
        for i in range(self.wage_table.columnCount()):
            self.wage_table_col_index[self.wage_table.horizontalHeaderItem(i).text()] = i
        self.export_main_table.triggered.connect(self.export_main_table_solt)
        self.export_user_table.triggered.connect(self.export_user_table_slot)
        self.export_main_table_result.triggered.connect(self.export_main_table_result_slot)
        self.export_wage_table.triggered.connect(self.export_wage_table_slot)
        self.center()
        self.main_table_path = None
        self.main_table_save_thread = threading.Thread()
        self.f_s_action.triggered.connect(self.f_s_action_slot)
        self.export_f_s_table.triggered.connect(self.export_f_s_table_slot)

        self.u_s_table_data_tmp = None
        self.u_s_action.triggered.connect(self.u_s_action_slot)
        self.export_u_s_table.triggered.connect(self.export_u_s_table_slot)
        self.tabWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tabWidget.customContextMenuRequested.connect(self.tabWidget_customContextMenuRequested_slot)
        self.function_tool_box = FunctionToolBox()
        self.function_tool_box.font_color.clicked.connect(self.table_font_color_slot)
        self.function_tool_box.fill_color.clicked.connect(self.table_fill_color_slot)

        self.username = None
        self.access_token = None

        self.http = Http()

    def adjust_main_table_widget(self):
        for k in self.col_index.keys():
            if k != "行程":
                self.main_table.resizeColumnToContents(self.col_index[k])
            else:
                self.main_table.setColumnWidth(self.col_index[k], 350)

    def adjust_main_table_result_widget(self):
        for k in self.col_index.keys():
            if k != "行程":
                self.main_table_result.resizeColumnToContents(self.col_index[k])
            else:
                self.main_table_result.setColumnWidth(self.col_index[k], 350)

    def adjust_user_table_widget(self):
        for k in self.user_table_col_index.keys():
            if k != "行程":
                self.user_table.resizeColumnToContents(self.user_table_col_index[k])
            else:
                self.user_table.setColumnWidth(self.user_table_col_index[k], 400)

    def adjust_wage_table_widget(self):
        for k in self.wage_table_col_index.keys():
            if k != "行程":
                self.wage_table.resizeColumnToContents(self.wage_table_col_index[k])
            else:
                self.wage_table.setColumnWidth(self.wage_table_col_index[k], 400)

    def table_font_color_slot(self):
        if self.tabWidget.currentIndex() == 0:
            table = self.main_table
            try:
                self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
            except Exception as e:
                pass
        elif self.tabWidget.currentIndex() == 1:
            table = self.main_table_result
            try:
                self.main_table_result.cellChanged.disconnect(self.main_table_result_cell_changed_slot)
            except Exception as e:
                pass
        elif self.tabWidget.currentIndex() == 2:
            table = self.user_table
        elif self.tabWidget.currentIndex() == 3:
            table = self.wage_table
        elif self.tabWidget.currentIndex() == 4:
            table = self.f_s_table
        elif self.tabWidget.currentIndex() == 5:
            table = self.u_s_table
        else:
            return None

        if table is None:
            return
        # a通道默认为255
        color = QtWidgets.QColorDialog.getColor(title="选择字体颜色")
        if not color.isValid():
            return
        select_range = table.selectedRanges()
        for select_area in select_range:
            top_row = select_area.topRow()
            bottom_row = select_area.bottomRow()
            left_column = select_area.leftColumn()
            right_column = select_area.rightColumn()
            for row in range(top_row, bottom_row + 1):
                for column in range(left_column, right_column + 1):
                    item = table.item(row, column)
                    if item is None:
                        continue
                    item.setForeground(color)
        if self.auto_save_checkbox.isChecked() and (table == self.main_table or table == self.main_table_result):
            self.save_main_table()
        if self.tabWidget.currentIndex() == 0:
            self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)
        elif self.tabWidget.currentIndex() == 1:
            self.main_table_result.cellChanged.connect(self.main_table_result_cell_changed_slot)

    def table_fill_color_slot(self):
        if self.tabWidget.currentIndex() == 0:
            table = self.main_table
            try:
                self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
            except Exception as e:
                pass
        elif self.tabWidget.currentIndex() == 1:
            table = self.main_table_result
            try:
                self.main_table_result.cellChanged.disconnect(self.main_table_result_cell_changed_slot)
            except Exception as e:
                pass
        elif self.tabWidget.currentIndex() == 2:
            table = self.user_table
        elif self.tabWidget.currentIndex() == 3:
            table = self.wage_table
        elif self.tabWidget.currentIndex() == 4:
            table = self.f_s_table
        elif self.tabWidget.currentIndex() == 5:
            table = self.u_s_table
        else:
            return None

        if table is None:
            return
        color = QtWidgets.QColorDialog.getColor(title="选择单元格填充颜色")
        if not color.isValid():
            return
        select_range = table.selectedRanges()
        for select_area in select_range:
            top_row = select_area.topRow()
            bottom_row = select_area.bottomRow()
            left_column = select_area.leftColumn()
            right_column = select_area.rightColumn()
            for row in range(top_row, bottom_row + 1):
                for column in range(left_column, right_column + 1):
                    item = table.item(row, column)
                    if item is None:
                        continue
                    item.setBackground(color)
        if self.auto_save_checkbox.isChecked() and (table == self.main_table or table == self.main_table_result):
            self.save_main_table()
        if self.tabWidget.currentIndex() == 0:
            self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)
        elif self.tabWidget.currentIndex() == 1:
            self.main_table_result.cellChanged.connect(self.main_table_result_cell_changed_slot)

    def tabWidget_customContextMenuRequested_slot(self, pos):
        # 不要窗口边缘
        # pos = QtCore.QPoint(pos.x(), pos.y() - 20)
        self.function_tool_box.move(self.mapToGlobal(pos))
        self.function_tool_box.hide()
        self.function_tool_box.show()

    def export_u_s_table_slot(self):
        if self.u_s_table.rowCount() == 0 or self.u_s_table_data_tmp is None:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "客户账单统计结果为空！",
                                                parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出财务统计表", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            self.statusbar.showMessage("用户取消！")
            return
        flag, info = export_u_s_table_widget(self.u_s_table, self.u_s_table_data_tmp, file_name)
        if flag:
            self.statusbar.showMessage("导出成功！")
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "导出成功！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
        else:
            self.statusbar.showMessage(info)

    def u_s_dialog_search_slot(self, target_year, yongchedanweis):
        # target_year yyyy
        # 查询yongchedanweis中的月度收入，支出，结余

        income_index = self.col_index['金额']
        expense_index = self.col_index['应付']
        balance_index = self.col_index['结余']
        yongchedanwei_index = self.col_index['用车单位']
        date_index = self.col_index['日期']
        data = {}
        # 每一个是{} key是月份，value是[income, expense, balance]
        for i in range(self.main_table.rowCount()):
            date = self.main_table.item(i, date_index).text()
            year = date.split('-')[0]
            if year != target_year:
                continue
            yongchedanwei = self.main_table.item(i, yongchedanwei_index).text()
            if yongchedanwei not in yongchedanweis:
                continue
            if yongchedanwei not in data:
                data[yongchedanwei] = {str(i): [0, 0, 0] for i in range(1, 13)}
            month = date.split('-')[1].lstrip('0')
            income = self.main_table.item(i, income_index).text()
            expense = self.main_table.item(i, expense_index).text()
            balance = self.main_table.item(i, balance_index).text()
            try:
                data[yongchedanwei][month][0] += float(income)
                data[yongchedanwei][month][1] += float(expense)
                data[yongchedanwei][month][2] += float(balance)
            except Exception as e:
                pass
        self.u_s_table_data_tmp = data
        self.u_s_table.setRowCount(12 * len(data))
        row_count = 0
        for yongchedanwei, value in data.items():
            for month, income_expense_balance in value.items():
                self.u_s_table.setItem(row_count, 0, QTableWidgetItem(yongchedanwei))
                self.u_s_table.setItem(row_count, 1, QTableWidgetItem(month))
                self.u_s_table.setItem(row_count, 2, QTableWidgetItem(str(income_expense_balance[0])))
                self.u_s_table.setItem(row_count, 3, QTableWidgetItem(str(income_expense_balance[1])))
                self.u_s_table.setItem(row_count, 4, QTableWidgetItem(str(income_expense_balance[2])))
                row_count += 1
        self.u_s_table.sortByColumn(0, Qt.AscendingOrder)
        self.tabWidget.setCurrentIndex(5)

    def u_s_action_slot(self):
        yongchedanweis = self.get_yongchedanwei_from_main_table()
        u_s_dialog = USDialog(yongchedanweis, parent=self)
        u_s_dialog.search_signal.connect(self.u_s_dialog_search_slot)
        u_s_dialog.exec_()

    def export_f_s_table_slot(self):
        if self.f_s_table.rowCount() == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "财务统计表为空！",
                                                parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出财务统计表", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            self.statusbar.showMessage("用户取消！")
            return
        flag, info = export_table_widget(self.f_s_table, file_name)
        if flag:
            self.statusbar.showMessage("导出成功！")
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "导出成功！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
        else:
            self.statusbar.showMessage(info)

    def f_s_dialog_search_slot(self, year, months):
        # year yyyy
        self.f_s_table.setColumnCount(3)
        income, income_idx = 0, self.col_index['金额']
        expense, expense_index = 0, self.col_index['应付']
        balance, balance_index = 0, self.col_index['结余']
        year_index = self.col_index['日期']
        months = sorted(months)
        if len(months) == 12:
            # 全年
            desc = [f'{year}年度总收入', f'{year}年度总支出', f'{year}年度结余']
        else:
            desc = [f'{months}月份总收入', f'{months}月份总支出', f'{months}月份总结余']
        self.f_s_table.setHorizontalHeaderLabels(desc)
        flag = False
        rows = []
        for i in range(self.main_table.rowCount()):
            year_table, month_table, _ = self.main_table.item(i, year_index).text().split('-')
            if year_table != year:
                # 判断年
                continue
            if len(months) != 12 and int(month_table) not in months:
                # 判断月
                continue
            income_str = self.main_table.item(i, income_idx).text().strip()
            expense_str = self.main_table.item(i, expense_index).text().strip()
            balance_str = self.main_table.item(i, balance_index).text().strip()
            try:
                income_number = float(income_str if income_str else '0')
                expense_number = float(expense_str if expense_str else '0')
                balance_number = float(balance_str if balance_str else '0')
            except Exception as e:
                rows.append(self.main_table.item(i, 0).text())
                income_number = 0
                expense_number = 0
                balance_number = 0
                flag = True
            income += income_number
            expense += expense_number
            balance += balance_number
        if flag:
            rows = sorted(rows)
            QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", f"序号为{rows}存在非纯数字情况！所在记录不进行计算！", parent=self).exec_()
        self.f_s_table.setRowCount(1)
        self.f_s_table.setItem(0, 0, QtWidgets.QTableWidgetItem(str(income)))
        self.f_s_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(expense)))
        self.f_s_table.setItem(0, 2, QtWidgets.QTableWidgetItem(str(balance)))
        self.f_s_table.resizeColumnsToContents()
        self.tabWidget.setCurrentIndex(4)

    def f_s_action_slot(self):
        f_s_dialog = FSDialog(parent=self)
        f_s_dialog.search_signal.connect(self.f_s_dialog_search_slot)
        f_s_dialog.exec_()

    def init_font(self):
        font = QApplication.font()
        self.setFont(font)

    def auto_save_checkbox_state_changed_slot(self, state):
        if state == Qt.Checked:
            # 触发一次保存
            self.save_main_table()
            self.statusbar.showMessage("自动保存已开启！")
        else:
            self.statusbar.showMessage("自动保存已关闭！")

    def wage_table_search_args_slot(self, start_date, end_date, driver, cheshudanwei, fu_state, shou_state):
        self.wage_table.clearContents()
        self.wage_table.setRowCount(0)
        for i in range(self.main_table.rowCount()):
            if start_date <= self.main_table.item(i, self.col_index["日期"]).text() <= end_date and \
                    self.main_table.item(i, self.col_index["司机"]).text() == driver and \
                    self.main_table.item(i, self.col_index["车属单位"]).text() == cheshudanwei:
                # 判断收付状态
                fu_ = self.main_table.item(i, self.col_index["付"]).text()
                shou_ = self.main_table.item(i, self.col_index["收"]).text()
                flag = True
                if flag and len(fu_state) != 0:
                    if fu_ not in fu_state:
                        flag = False
                if flag and len(shou_state) != 0:
                    if shou_ not in shou_state:
                        flag = False
                if flag:
                    self.wage_table.insertRow(self.wage_table.rowCount())
                    for key in self.wage_table_col_index:
                        self.wage_table.setItem(self.wage_table.rowCount() - 1, self.wage_table_col_index[key],
                                                self.main_table.item(i, self.col_index[key]).clone())

        self.adjust_wage_table_widget()
        self.tabWidget.setCurrentIndex(3)
        self.statusbar.showMessage("生成工资表成功！")

    def create_wage_table_btn_slot(self):
        # 传入所有的 “司机” “车属单位”
        driver_set = self.get_driver_from_main_table()
        cheshudanwei_set = self.get_cheshudanwei_from_main_table()
        fu_set = self.get_fu_state_from_main_table()
        shou_set = self.get_shou_state_from_main_table()
        create_wage_table_dialog = WageTableCreateDialog(driver_set, cheshudanwei_set, fu_set, shou_set, parent=self)
        create_wage_table_dialog.search_signal.connect(self.wage_table_search_args_slot)
        create_wage_table_dialog.exec_()

    def export_main_table_result_slot(self):
        if self.main_table_result.rowCount() == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "明细表查询结果为空！",
                                                parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出明细表查询结果", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            self.statusbar.showMessage("用户取消！")
            return
        # 保存
        flag, info = export_table_widget(self.main_table_result, file_name)
        if flag:
            self.export_table_success()
        else:
            self.export_table_failed(info)

    def export_wage_table_slot(self):
        if self.wage_table.rowCount() == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "工资表为空！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出工资表", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            self.statusbar.showMessage("用户取消！")
            return
        # 保存工资表
        total_money = 0
        for i in range(self.wage_table.rowCount()):
            total_money += float(self.wage_table.item(i, self.wage_table_col_index["应付"]).text())
        flag, info = export_table_widget_with_money(self.wage_table, file_name, "总应付", total_money)
        if flag:
            self.export_table_success()
        else:
            self.export_table_failed(info)

    def export_table_success(self):
        message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "导出成功！", parent=self)
        message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
        message_box.exec_()
        self.statusbar.showMessage("导出成功！")

    def export_table_failed(self, info):
        message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "导出失败！", parent=self)
        message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
        message_box.exec_()
        self.statusbar.showMessage(info)

    def export_main_table_solt(self):
        if self.main_table.rowCount() == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "明细表为空！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出明细表", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            return
        # 保存
        flag, info = export_table_widget(self.main_table, file_name)
        if flag:
            self.export_table_success()
        else:
            self.export_table_failed(info)

    def export_user_table_slot(self):
        if self.user_table.rowCount() == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "客户账单为空！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        file_name, file_type = QFileDialog.getSaveFileName(self, "导出用户表", "", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            return
        # 保存客户表
        total_money = 0
        for i in range(self.user_table.rowCount()):
            total_money += float(self.user_table.item(i, self.user_table_col_index["金额"]).text())
        export_table_widget_with_money(self.user_table, file_name, "总金额", total_money)
        self.export_table_success()

    def get_yongchedanwei_from_main_table(self):
        # 获取“用车单位”列表
        yongchedanwei_set = set()
        for i in range(self.main_table.rowCount()):
            yongchedanwei_set.add(self.main_table.item(i, self.col_index["用车单位"]).text())
        yongchedanwei_set = sorted(yongchedanwei_set)
        return yongchedanwei_set

    def get_cheshudanwei_from_main_table(self):
        # 获取“车属单位”列表
        cheshudanwei_set = set()
        for i in range(self.main_table.rowCount()):
            cheshudanwei_set.add(self.main_table.item(i, self.col_index["车属单位"]).text())
        cheshudanwei_set = sorted(cheshudanwei_set)
        return cheshudanwei_set

    def get_driver_from_main_table(self):
        # 获取“司机”列表
        driver_set = set()
        for i in range(self.main_table.rowCount()):
            driver_set.add(self.main_table.item(i, self.col_index["司机"]).text())
        driver_set = sorted(driver_set)
        return driver_set

    def get_fu_state_from_main_table(self):
        # 获取 "付" 列表
        fu_state_set = set()
        for i in range(self.main_table.rowCount()):
            fu_state_set.add(self.main_table.item(i, self.col_index["付"]).text())
        fu_state_set = sorted(fu_state_set)
        return fu_state_set

    def get_shou_state_from_main_table(self):
        # 获取 "收" 列表
        shou_state_set = set()
        for i in range(self.main_table.rowCount()):
            shou_state_set.add(self.main_table.item(i, self.col_index["收"]).text())
        shou_state_set = sorted(shou_state_set)
        return shou_state_set

    def create_user_table_btn_slot(self):
        # 打开用户表创建窗口
        yongchedanweis = self.get_yongchedanwei_from_main_table()
        fus = self.get_fu_state_from_main_table()
        shous = self.get_shou_state_from_main_table()
        user_table_create_dialog = UserTableCreateDialog(yongchedanweis, fus, shous, parent=self)
        user_table_create_dialog.create_user_table_args_signal.connect(self.create_user_table_args_slot)
        user_table_create_dialog.exec_()

    def create_user_table_args_slot(self, start_date, end_date, checked_checkbox_text_set, fu_state, shou_state):
        # 查找main_table中符合的行到user_table
        # start_date: yyyy-MM-dd
        # end_date: yyyy-MM-dd
        # checked_checkbox_text_set: 选中的checkbox的text列表为"用车单位"
        self.user_table.clearContents()
        self.user_table.setRowCount(0)
        for i in range(self.main_table.rowCount()):
            date = self.main_table.item(i, self.col_index["日期"]).text()
            if start_date <= date <= end_date:
                yongchedanwei = self.main_table.item(i, self.col_index["用车单位"]).text()
                if yongchedanwei in checked_checkbox_text_set:
                    # 判断付收状态
                    fu_ = self.main_table.item(i, self.col_index["付"]).text()
                    shou_ = self.main_table.item(i, self.col_index["收"]).text()
                    flag = True
                    if flag and len(fu_state) != 0:
                        if fu_ not in fu_state:
                            flag = False
                    if flag and len(shou_state) != 0:
                        if shou_ not in shou_state:
                            flag = False
                    if flag:
                        self.user_table.insertRow(self.user_table.rowCount())
                        for k in self.user_table_col_index:
                            main_table_col = self.col_index[k]
                            user_table_col = self.user_table_col_index[k]
                            content_itme = self.main_table.item(i, main_table_col)
                            self.user_table.setItem(self.user_table.rowCount() - 1, user_table_col,
                                                    content_itme.clone())

        self.adjust_user_table_widget()
        self.tabWidget.setCurrentIndex(2)
        self.statusbar.showMessage("生成客户账单成功！")

    def save_main_table(self):
        print('a')
        # 保存总表
        # 保存到self.main_table_file_name
        # 保存到self.main_table
        if self.auto_save_checkbox.isChecked() and self.main_table_path is not None:
            # 放在线程中执行
            if self.main_table_save_thread.is_alive():
                self.main_table_save_thread.join()
            self.main_table_save_thread = threading.Thread(target=export_table_widget,
                                                           args=[self.main_table, self.main_table_path])
            self.main_table_save_thread.start()
            self.statusbar.showMessage("保存成功！")
        self.adjust_main_table_widget()

    def center(self):
        # 居中显示 且占据屏幕70%
        screen = QDesktopWidget().screenGeometry()
        self.resize(int(screen.width() * 0.7), int(screen.height() * 0.7))
        size = self.geometry()
        self.move((screen.width() - size.width()) // 2, (screen.height() - size.height()) // 2)

    def get_main_table_max_index(self):
        # 获取“序号”这一列的最大值
        # 根据表头获取“序号”所在的列
        index_col = self.col_index.get("序号", -1)
        if index_col == -1:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "序号列不存在！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return
        max_index = 0
        for i in range(self.main_table.rowCount()):
            item = self.main_table.item(i, index_col).text()
            if item.isdigit():
                max_index = max(max_index, int(item))
        return max_index

    def main_table_add_event(self, info):
        # 添加一行
        # date: yyyy-MM-dd (str)
        date = info["date"]
        driver = info["driver"]
        car_type = info["car_type"]
        car_number = info["car_number"]
        qiandan = info["qiandan"]
        cheshudanwei = info["cheshudanwei"]
        yongchedanwei = info["yongchedanwei"]
        time = info["time"]
        money = info["money"]
        shou = info["shou"]
        piao = ""
        fu = info["fu"]
        pay = info["pay"]
        try:
            # 如果是整数，就转换成整数
            balance = float(money) - float(pay)
            if balance == int(balance):
                balance = int(balance)
            balance = str(balance)
        except Exception as e:
            balance = ""
        remark = info["remark"]
        itinerary = info["itinerary"]
        cur_index = self.get_main_table_max_index() + 1
        try:
            self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
        except Exception as e:
            pass
        self.main_table.insertRow(self.main_table.rowCount())
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["序号"],
                                QtWidgets.QTableWidgetItem(str(cur_index)))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["日期"],
                                QtWidgets.QTableWidgetItem(date))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["司机"],
                                QtWidgets.QTableWidgetItem(driver))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["车号"],
                                QtWidgets.QTableWidgetItem(car_number))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["车型"],
                                QtWidgets.QTableWidgetItem(car_type))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["签单"],
                                QtWidgets.QTableWidgetItem(qiandan))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["车属单位"],
                                QtWidgets.QTableWidgetItem(cheshudanwei))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["用车单位"],
                                QtWidgets.QTableWidgetItem(yongchedanwei))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["时间"],
                                QtWidgets.QTableWidgetItem(time))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["行程"],
                                QtWidgets.QTableWidgetItem(itinerary))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["金额"],
                                QtWidgets.QTableWidgetItem(money))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["收"], QtWidgets.QTableWidgetItem(shou))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["票"], QtWidgets.QTableWidgetItem(piao))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["应付"], QtWidgets.QTableWidgetItem(pay))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["付"], QtWidgets.QTableWidgetItem(fu))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["结余"],
                                QtWidgets.QTableWidgetItem(balance))
        self.main_table.setItem(self.main_table.rowCount() - 1, self.col_index["备注"],
                                QtWidgets.QTableWidgetItem(remark))
        self.add_completed.emit()
        self.main_table.sortItems(self.col_index["日期"], Qt.AscendingOrder)
        self.adjust_main_table_widget()
        # main_table滑动到最后一行
        self.main_table.scrollToBottom()
        # 重置序号
        self.reset_main_table_index()
        self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)
        if self.auto_save_checkbox.isChecked():
            self.save_main_table()

    def reset_main_table_index(self):
        # 重置序号
        for i in range(self.main_table.rowCount()):
            self.main_table.item(i, self.col_index["序号"]).setText(str(i + 1))

    def main_table_del_btn_slot(self):
        # 删除选中的行
        selected_rows = self.main_table.selectedItems()
        if len(selected_rows) == 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Warning, "警告", "请先选中要删除的行！",
                                                parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.exec_()
            return

        message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示", "确定删除选中的行？", parent=self)
        message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
        message_box.addButton(self.tr("取消"), QtWidgets.QMessageBox.NoRole)
        ret = message_box.exec_()
        if ret == 1:
            return
        selected_rows = list(set([item.row() for item in selected_rows]))
        selected_rows.sort(reverse=True)
        for row in selected_rows:
            self.main_table.removeRow(row)
        if self.auto_save_checkbox.isChecked():
            self.save_main_table()

    def main_table_add_slot(self):
        main_table_add_dialog = MainTableAddDialog(parent=self)
        main_table_add_dialog.add_confirm.connect(self.main_table_add_event)
        self.add_completed.connect(main_table_add_dialog.add_completed_slot)
        main_table_add_dialog.exec_()

    def search_table_from_args(self, args):
        try:
            self.main_table_result.cellChanged.disconnect(self.main_table_result_cell_changed_slot)
        except Exception as e:
            pass
        # 假设序号是唯一的，不可以修改
        # 清空表格 包括headers
        self.main_table_result.setRowCount(0)
        date_start = args["date_start"]
        date_end = args["date_end"]
        yongchedanwei_set = args["yongchedanweis"]
        cheshudanwei_set = args["cheshudanweis"]
        driver_set = args["drivers"]
        itinerary = args["itinerary"]
        fu = args["fu"]
        shou = args["shou"]
        for i in range(self.main_table.rowCount()):
            date = self.main_table.item(i, self.col_index["日期"]).text()
            yongchedanwei_ = self.main_table.item(i, self.col_index["用车单位"]).text()
            cheshudanwei_ = self.main_table.item(i, self.col_index["车属单位"]).text()
            driver_ = self.main_table.item(i, self.col_index["司机"]).text()
            itinerary_ = self.main_table.item(i, self.col_index["行程"]).text()
            fu_ = self.main_table.item(i, self.col_index["付"]).text()
            shou_ = self.main_table.item(i, self.col_index["收"]).text()
            flag = True
            if not (date_start <= date <= date_end):
                flag = False
            if flag and len(yongchedanwei_set) != 0:
                if yongchedanwei_ not in yongchedanwei_set:
                    flag = False
            if flag and len(cheshudanwei_set) != 0:
                if cheshudanwei_ not in cheshudanwei_set:
                    flag = False
            if flag and len(driver_set) != 0:
                if driver_ not in driver_set:
                    flag = False
            if flag and itinerary not in itinerary_:
                flag = False

            if flag and len(fu) != 0:
                if fu_ not in fu:
                    flag = False
            if flag and len(shou) != 0:
                if shou_ not in shou:
                    flag = False
            if flag:
                self.main_table_result.insertRow(self.main_table_result.rowCount())
                for j in range(self.main_table_result.columnCount()):
                    self.main_table_result.setItem(self.main_table_result.rowCount() - 1, j,
                                                   self.main_table.item(i, j).clone())
        self.adjust_main_table_result_widget()
        self.main_table_result.cellChanged.connect(self.main_table_result_cell_changed_slot)
        self.tabWidget.setCurrentIndex(1)

    def main_table_search_slot(self):
        yongchedanweis = self.get_yongchedanwei_from_main_table()
        cheshudanweis = self.get_cheshudanwei_from_main_table()
        drivers = self.get_driver_from_main_table()
        fus = self.get_fu_state_from_main_table()
        shous = self.get_shou_state_from_main_table()
        search_dialog = MainTableSearchWidget(yongchedanweis, cheshudanweis, drivers, fus, shous, parent=self)
        search_dialog.search_args_signal.connect(self.search_table_from_args)
        search_dialog.exec_()

    def open_table_main_slot(self):
        try:
            self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
        except Exception as e:
            pass
        # 打开Excel类的文件 包括xlsm
        self.statusbar.showMessage("用户选择中。")
        file_name, file_type = QFileDialog.getOpenFileName(self, "选取文件", "./", "Excel(*.xlsx *.xls *.xlsm)")
        if file_name == "":
            self.statusbar.showMessage("用户取消！")
            return
        try:
            sheet_names = get_sheet_names(file_name)
            self.statusbar.showMessage("选择表格中。")
            if len(sheet_names) == 0:
                self.statusbar.showMessage("打开失败！")
                return
            elif len(sheet_names) == 1:
                sheet_name = sheet_names[0]
            else:
                # 只用一次就释放了
                input_dialog = QtWidgets.QInputDialog(parent=self)
                input_dialog.setOkButtonText('确认')
                input_dialog.setCancelButtonText('取消')
                input_dialog.setLabelText('请选择表格')
                input_dialog.setComboBoxItems(sheet_names)
                input_dialog.setComboBoxEditable(False)
                input_dialog.setWindowTitle('选择表格')
                ok = input_dialog.exec_()
                sheet_name = input_dialog.textValue()
                if not ok:
                    self.statusbar.showMessage("用户取消！")
                    return
            self.statusbar.showMessage("打开中。")
            # 使用openpyxl读取excel
            excel_table = openpyxl.load_workbook(file_name)[sheet_name]
            data_row_count = excel_table.max_row - 1
            self.main_table.setRowCount(data_row_count)
            self.main_table.setColumnCount(excel_table.max_column)
            self.main_table_result.setColumnCount(excel_table.max_column)
            # 读取数据
            for i in range(data_row_count):
                for j in range(excel_table.max_column):
                    cell_item = excel_table.cell(row=i + 2, column=j + 1)
                    background_color = "FFFFFF"
                    foreground_color = "000000"
                    if cell_item.value is None:
                        item_content = ""
                    else:
                        try:
                            background_color = cell_item.fill.fgColor.rgb[2:].upper()
                            foreground_color = cell_item.font.color.rgb[2:].upper()
                        except Exception as e:
                            background_color = "FFFFFF"
                            foreground_color = "000000"
                        if self.col_index['日期'] == j and isinstance(cell_item.value, datetime.datetime):
                            item_content = cell_item.value.strftime("%Y-%m-%d")
                        else:
                            item_content = str(cell_item.value)
                    item = QTableWidgetItem(item_content.strip())
                    item.setBackground(QColor("#" + background_color))
                    if item_content != "":
                        item.setForeground(QColor("#" + foreground_color))
                    self.main_table.setItem(i, j, item)
            self.cal_balance_on_main_table()
            self.adjust_main_table_widget()
            self.statusbar.showMessage("打开成功！")
            self.main_table_path = file_name
            self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)
        except Exception as e:
            print(e)
            self.statusbar.showMessage("打开失败！")
        self.tabWidget.setCurrentIndex(0)

    def cal_balance(self, table, row) -> str:
        money = table.item(row, self.col_index["金额"]).text()
        pay = table.item(row, self.col_index["应付"]).text()
        try:
            balance = float(money) - float(pay)
            if balance == int(balance):
                balance = int(balance)
            balance = str(balance)
        except Exception as e:
            balance = ""
        return balance

    def cal_balance_on_main_table(self):
        for i in range(self.main_table.rowCount()):
            balance = self.cal_balance(self.main_table, i)
            self.main_table.setItem(i, self.col_index["结余"], QTableWidgetItem(balance))

    def reset_main_table_according_to_data(self):
        try:
            self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
        except Exception as e:
            pass
        self.main_table.sortByColumn(self.col_index["日期"], QtCore.Qt.AscendingOrder)
        self.reset_main_table_index()
        self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)

    def main_table_cell_changed_slot(self, row, col):
        print(row, col)
        # 结余自动计算
        if col == self.col_index["金额"] or col == self.col_index["应付"]:
            balance = self.cal_balance(self.main_table, row)
            # 不要触发cellChanged
            try:
                self.main_table.cellChanged.disconnect(self.main_table_cell_changed_slot)
            except Exception as e:
                pass
            self.main_table.setItem(row, self.col_index["结余"], QtWidgets.QTableWidgetItem(balance))
            self.main_table.cellChanged.connect(self.main_table_cell_changed_slot)
        elif col == self.col_index["日期"]:
            self.reset_main_table_according_to_data()
        if self.auto_save_checkbox.isChecked():
            self.save_main_table()
        self.adjust_main_table_widget()

    def main_table_result_cell_changed_slot(self, row, col):
        # 同步修改到main_table
        # 以row这一样的“序号”列为主键
        index = self.main_table_result.item(row, self.col_index["序号"]).text()
        for i in range(self.main_table.rowCount()):
            if self.main_table.item(i, self.col_index["序号"]).text() == index:
                self.main_table.item(i, col).setText(self.main_table_result.item(row, col).text())
                break
        # 如果是"金额" "应付"列 要同时修改 "结余"列 由于main_table已经设置了changed监听 所以会自动修改
        # 所以这里只需要修改main_table_result
        if col == self.col_index["金额"] or col == self.col_index["应付"]:
            balance = self.cal_balance(self.main_table_result, row)
            try:
                self.main_table_result.cellChanged.disconnect(self.main_table_result_cell_changed_slot)
            except Exception as e:
                pass
            self.main_table_result.setItem(row, self.col_index["结余"], QtWidgets.QTableWidgetItem(balance))
            self.main_table_result.cellChanged.connect(self.main_table_result_cell_changed_slot)
        self.adjust_main_table_result_widget()

    def closeEvent(self, e: QtGui.QCloseEvent) -> None:
        if self.main_table.rowCount() != 0 or self.main_table_result.rowCount() != 0 \
                or self.user_table.rowCount() != 0 or self.wage_table.rowCount() != 0:
            message_box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, "提示",
                                                "退出之前请确保所有数据文件已导出！", parent=self)
            message_box.addButton(self.tr("确定"), QtWidgets.QMessageBox.YesRole)
            message_box.addButton(self.tr("取消"), QtWidgets.QMessageBox.NoRole)
            ret = message_box.exec_()
            if ret == 1:
                e.ignore()
                return
        e.accept()
        self.function_tool_box.close()

    def resizeEvent(self, a0: QtGui.QResizeEvent) -> None:
        self.adjust_main_table_widget()
        self.adjust_main_table_result_widget()
        self.adjust_user_table_widget()
        self.adjust_wage_table_widget()
